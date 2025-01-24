import random
import openpyxl
import matplotlib.pyplot as grafico
from estadisticas import solucion, nivel_dificultad

def partida_modosolitario():
    print("Modo Solitario")
    #Pedimos a la consola que nos genere un numero aleatorio entre 1 y 1000.
    numero_adivinar=random.randint(1,1000)
    
        
    #Le pedimos al jugador que nos indique el nivel de dificultad, para avisarle de los intentos que dispone   
    intentos=nivel_dificultad()
    print("Dispones de",intentos, "intentos para adivinar el numero")

    #El jugador debe adivinar el numero y se le iran dando pistas 
    for intentos in range(1,intentos+1):
        adivinanza=int(input("Intento numero " +str(intentos)+ ", Introduce tu numero:"))

        #Nos aseguramos que dicho numero esta entre 1 y 1000
        if adivinanza>1000 or adivinanza<1:
            print("Por favor introduce un numero entre 1 y 1000:")

        if adivinanza<numero_adivinar:
            print("El numero a adivinar es mayor")
        elif adivinanza>numero_adivinar:
            print("El numero a adivinar es menor")
        else:
            print("Enhorabuena, has acertado el numero")
            nombre=input("¿Cual es tu nombre?")
            solucion(nombre, "Ha ganado")
            return
    print("Has superado el numero de intentos, el numero era:" +str(numero_adivinar))
    #Se le pide al jugador el nombre y se guarda su resultado
    nombre=input("¿Cual es tu nombre?")
    solucion (nombre, "Ha perdido")


def partida_dos_jugadores():
    print("Modo 2 jugadores")
    #Le pedimos al jugador 1 que nos indique el numero que el otro jugador debe adivinar
    numero_adivinar=int(input("Jugador 1 introduce un numero:"))

    #Nos aseguramos que el numero que introduce esta entre 1 y 1000.
    while numero_adivinar>1000 or numero_adivinar<1:
        numero_adivinar=int(input("Por favor introduce un numero entre 1 y 1000:"))

    #Le pedimos al jugador que nos indique el nivel de dificultad, para avisarle de los intentos que dispone  
    intentos=nivel_dificultad()
    print("Dispones de" , intentos, "intentos para adivinar el numero")
    
   #El jugador debe adivinar el numero y se le iran dando pista
    for intentos in range(1,intentos+1):
        adivinanza=int(input("Intento numero " +str(intentos)+ ",Introduce tu numero"))
        
        #Nos aseguramos que el numero que introduce esta entre 1 y 1000.
        if adivinanza>1000 or adivinanza<1:
            print("Por favor introduce un numero entre 1 y 1000:")
            
        if adivinanza>numero_adivinar:
            print("El numero a adivinar es menor")
        elif adivinanza<numero_adivinar:
            print("El numero a adivinar es mayor")
        else:
            print("Enhorabuena, has acertado el numero")
            nombre=input("¿Cual es tu nombre?")
            solucion (nombre, "Ha ganado")
            return
    print("Has superado el numero de intentos, el numero era:" + str(numero_adivinar))
    #Se le pide al jugador el nombre y se guarda su resultado
    nombre=input("¿Cual es tu nombre?")
    solucion (nombre, "Ha perdido")     

def estadistica(libro):
    print("Estadistica de los jugadores")
    #Indicamos la ruta del archivo donde se van a ir almacenando los resultados del juego
    archivo= ("C:\\Users\\celia\\OneDrive\\Escritorio\\DATA SCIENCE\\Programacion Python\\datos_estadisticas.xlsx")
    
    #Cargamos nuestro fichero Excel
    libro=openpyxl.load_workbook("C:\\Users\\celia\\OneDrive\\Escritorio\\DATA SCIENCE\\Programacion Python\\datos_estadisticas.xlsx")

    #Cargamos nuesta hoja
    hoja_estadisticas=libro["hoja_estadisticas"]

    #Creamos la lista para poder  mostrar nuestros resultados
    resultados=[]

    #Extraemos los datos de cada hoja
    for fila in range(2,hoja_estadisticas.max_row +1):
        
        #Agregamos los valores 
        cell_value = hoja_estadisticas.cell(row=fila, column=2).value  
        if cell_value is not None:
            resultados.append(cell_value)
            
    #Contamos los que ganan y los que pierden
    ganados=0
    perdidos=0

    for resultado in resultados:
        if resultado=="Ha ganado":
            ganados+=1
        else: 
            perdidos +=1
    
    #creamos el gráfico    
    grafico.bar (["ganados", "perdidos"], [ganados, perdidos],color=["green","red"])
    grafico.ylabel ("Nº de personas")
    grafico.title("Ganadores vs perdedores")
    grafico.show()
            
    #Guardamos los datos
    libro.save(archivo)
    print("Los datos anteriores se han guardado correctamente")
