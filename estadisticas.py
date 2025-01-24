import openpyxl
import matplotlib.pyplot as grafico

#Definimos una funcion, donde vamos a ir guardando los resultados del juego
def solucion(nombre,resultado):
    archivo=("C:\\Users\\celia\\OneDrive\\Escritorio\\DATA SCIENCE\\Programacion Python\\datos_estadisticas.xlsx")
    #Cargamos nuestro fichero    
    try:
        libro=openpyxl.load_workbook(archivo)
        
    except FileNotFoundError:
        print("No se encuentra el fichero")
        return
   #Nos aseguramos que existe la hoja
    if "hoja_estadisticas" not in libro.sheetnames:
        print("No se encuentra la hoja")
        return

    hoja_estadisticas=libro["hoja_estadisticas"]

    #Nos aseguramos que van a salir estos encabezados al comenzar a rellenar nuestro fichero, cuando este vacio
    if hoja_estadisticas.max_row==1 and hoja_estadisticas["A1"].value != "Nombre": 
        hoja_estadisticas.append (["Nombre","Resultado"])
        

    if nombre=="":
       nombre="Desconocido"
        
    #Guardamos los datos en la hoja especifica
     
    hoja_estadisticas.append([nombre,resultado]) 
    
        
    #Guardamos los datos en nuestro archivo   
    libro.save(archivo)
    print("Resultados guardados correctamente")

def nivel_dificultad():
    #Nos aseguramos que el individuo introduzca una opcion valida
    while True:
        print("Elige nivel de dificultad: \n 1.Facil(20 intentos) \n 2.Medio(12 intentos) \n 3.Dificil(5 intentos)")
        
        
        #Pedimos al jugador que nos indique el nivel de dificultad con el que quiere jugar
        dificultad=int(input("Introduce el nivel de dificultad entre 1 y 3:"))
        if dificultad==1:
            return 20
        elif dificultad==2:
            return 12
        elif dificultad==3:
            return 5
        else:
            print("Nivel de dificultad no valido.Intentelo de nuevo")
            